from pathlib import Path

import pytest
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl import load_workbook
from openpyxl import Workbook

from harmony_translate.config import AppConfig
from harmony_translate.cli import build_parser
from harmony_translate.pipeline import run_pipeline


def build_sample_workbook() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = "Sheet1"
    worksheet["A13"] = "SHOT CODE"
    worksheet["P13"] = "ANIMATION"
    worksheet["Y13"] = "NOTES"
    worksheet["A14"] = "HH0304_010_0010"
    worksheet["P14"] = "No character Anim. Add eases to camera move."
    worksheet["Y14"] = "No character Anim. Add eases to camera move."
    worksheet["A15"] = "HH0304_010_0020"
    worksheet["P15"] = "Natural acting note for rough animation."
    worksheet["Y15"] = "Natural acting note for rough animation."
    worksheet["A16"] = "HH0304_010_0030"
    worksheet["P16"] = "Pickup this section with limited animation."
    worksheet["Y16"] = "Pickup this section with limited animation."
    worksheet["A17"] = "HH0304_010_0040"
    worksheet["P17"] = "Storybook style shot, mostly tweens."
    worksheet["Y17"] = "Storybook style shot, mostly tweens."
    worksheet["A18"] = "HH0304_010_0050"
    worksheet["P18"] = "Hold the pose longer before the turn."
    worksheet["Y18"] = "Hold the pose longer before the turn."
    return workbook


def test_pipeline_runs_in_preview_mode_without_api_key(tmp_path: Path) -> None:
    workbook = build_sample_workbook()
    input_path = tmp_path / "sample.xlsx"
    workbook.save(input_path)

    config = AppConfig(
        input_path=input_path,
        output_dir=tmp_path / "out",
        sheet_name=None,
        selected_columns=[],
        preserve_original_sheet=False,
        mapped_cell_mode="translation_only",
        glossary_path=Path("/Users/usabatch/coding/hazbin_project/glossary.tsv"),
        exclude_patterns_path=Path(
            "/Users/usabatch/coding/hazbin_project/exclude_patterns.yaml"
        ),
        target_lang="KO",
        source_lang="EN",
        provider="deepl",
        deepl_api_key="",
        deepl_base_url="https://api-free.deepl.com",
        gemini_api_key="",
        gemini_model="gemini-2.0-flash",
        gemini_base_url="https://generativelanguage.googleapis.com",
        cache_path=tmp_path / "cache.sqlite3",
    )

    result = run_pipeline(config)

    assert result.preview_mode is True
    assert result.translated_path.exists()
    assert result.source_mapped_path.exists()
    assert result.source_mapped_path.name == "sample_KO.xlsx"
    assert result.audit_path.exists()
    assert result.usage_path.exists()


def test_cli_parser_defaults_input_to_sample_path() -> None:
    parser = build_parser()
    args = parser.parse_args([])
    assert str(args.input).endswith("HH0304-Episodic_Lead_Sheet_LIVE_Yeson.xlsx")
    assert args.provider == "gemini"
    assert args.gemini_model == "gemini-2.5-flash"


def test_pipeline_translates_only_unique_missing_texts(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    workbook = build_sample_workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["P15"] = "Repeat me"
    worksheet["Y15"] = "Repeat me"
    worksheet["P16"] = "Repeat me"
    worksheet["Y16"] = "Repeat me"
    input_path = tmp_path / "sample.xlsx"
    workbook.save(input_path)

    calls: list[list[str]] = []

    class FakeDeepLClient:
        def __init__(self, api_key: str, base_url: str) -> None:
            self.api_key = api_key
            self.base_url = base_url

        def usage(self):
            class Usage:
                character_count = 0
                character_limit = 500000

            return Usage()

        def translate_batch(
            self, texts: list[str], *, source_lang: str, target_lang: str
        ) -> list[str]:
            calls.append(list(texts))
            return [f"KO:{text}" for text in texts]

    monkeypatch.setattr("harmony_translate.pipeline.DeepLClient", FakeDeepLClient)

    config = AppConfig(
        input_path=input_path,
        output_dir=tmp_path / "out",
        sheet_name=None,
        selected_columns=[],
        preserve_original_sheet=False,
        mapped_cell_mode="translation_only",
        glossary_path=Path("/Users/usabatch/coding/hazbin_project/glossary.tsv"),
        exclude_patterns_path=Path(
            "/Users/usabatch/coding/hazbin_project/exclude_patterns.yaml"
        ),
        target_lang="KO",
        source_lang="EN",
        provider="deepl",
        deepl_api_key="test-key",
        deepl_base_url="https://api-free.deepl.com",
        gemini_api_key="",
        gemini_model="gemini-2.0-flash",
        gemini_base_url="https://generativelanguage.googleapis.com",
        cache_path=tmp_path / "cache.sqlite3",
    )

    run_pipeline(config)

    flattened_calls = [text for batch in calls for text in batch]
    assert flattened_calls.count("Repeat me") == 1


def test_pipeline_does_not_preblock_when_usage_near_limit(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    workbook = build_sample_workbook()
    input_path = tmp_path / "sample.xlsx"
    workbook.save(input_path)

    class FakeDeepLClient:
        def __init__(self, api_key: str, base_url: str) -> None:
            self.api_key = api_key
            self.base_url = base_url

        def usage(self):
            class Usage:
                character_count = 499990
                character_limit = 500000

            return Usage()

        call_count = 0

        def translate_batch(
            self, texts: list[str], *, source_lang: str, target_lang: str
        ) -> list[str]:
            FakeDeepLClient.call_count += 1
            return [f"KO:{text}" for text in texts]

    monkeypatch.setattr("harmony_translate.pipeline.DeepLClient", FakeDeepLClient)

    config = AppConfig(
        input_path=input_path,
        output_dir=tmp_path / "out",
        sheet_name=None,
        selected_columns=[],
        preserve_original_sheet=False,
        mapped_cell_mode="translation_only",
        glossary_path=Path("/Users/usabatch/coding/hazbin_project/glossary.tsv"),
        exclude_patterns_path=Path(
            "/Users/usabatch/coding/hazbin_project/exclude_patterns.yaml"
        ),
        target_lang="KO",
        source_lang="EN",
        provider="deepl",
        deepl_api_key="test-key",
        deepl_base_url="https://api-free.deepl.com",
        gemini_api_key="",
        gemini_model="gemini-2.0-flash",
        gemini_base_url="https://generativelanguage.googleapis.com",
        cache_path=tmp_path / "cache.sqlite3",
    )

    result = run_pipeline(config)
    assert result.translated_path.exists()
    assert FakeDeepLClient.call_count > 0


def test_pipeline_writes_translations_back_to_source_mapped_workbook(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    workbook = build_sample_workbook()
    input_path = tmp_path / "sample.xlsx"
    workbook.save(input_path)

    class FakeDeepLClient:
        def __init__(self, api_key: str, base_url: str) -> None:
            self.api_key = api_key
            self.base_url = base_url

        def usage(self):
            class Usage:
                character_count = 0
                character_limit = 500000

            return Usage()

        def translate_batch(
            self, texts: list[str], *, source_lang: str, target_lang: str
        ) -> list[str]:
            return [f"KO:{text}" for text in texts]

    monkeypatch.setattr("harmony_translate.pipeline.DeepLClient", FakeDeepLClient)

    config = AppConfig(
        input_path=input_path,
        output_dir=tmp_path / "out",
        sheet_name=None,
        selected_columns=[],
        preserve_original_sheet=True,
        mapped_cell_mode="translation_only",
        glossary_path=Path("/Users/usabatch/coding/hazbin_project/glossary.tsv"),
        exclude_patterns_path=Path(
            "/Users/usabatch/coding/hazbin_project/exclude_patterns.yaml"
        ),
        target_lang="KO",
        source_lang="EN",
        provider="deepl",
        deepl_api_key="test-key",
        deepl_base_url="https://api-free.deepl.com",
        gemini_api_key="",
        gemini_model="gemini-2.0-flash",
        gemini_base_url="https://generativelanguage.googleapis.com",
        cache_path=tmp_path / "cache.sqlite3",
    )

    result = run_pipeline(config)
    translated = load_workbook(result.translated_path)
    translated_sheet = translated["Sheet1"]
    mapped = load_workbook(result.source_mapped_path)
    mapped_sheet = mapped["Sheet1"]

    assert (
        mapped_sheet["P14"].value == "KO:No character Anim. Add eases to camera move."
    )
    assert translated_sheet["AA14"].font.color is not None
    assert translated_sheet["AA14"].font.color.rgb == "FF0000FF"
    assert mapped_sheet["P14"].font.color is not None
    assert mapped_sheet["P14"].font.color.rgb == "FF0000FF"
    assert (
        mapped_sheet["Y14"].value == "KO:No character Anim. Add eases to camera move."
    )
    original_sheet = mapped["Sheet1_ORIGINAL"]
    assert original_sheet["P14"].value == "No character Anim. Add eases to camera move."
    assert original_sheet["Y14"].value == "No character Anim. Add eases to camera move."


def test_pipeline_can_show_original_and_translation_in_same_cell(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    workbook = build_sample_workbook()
    input_path = tmp_path / "sample.xlsx"
    workbook.save(input_path)

    class FakeDeepLClient:
        def __init__(self, api_key: str, base_url: str) -> None:
            self.api_key = api_key
            self.base_url = base_url

        def usage(self):
            class Usage:
                character_count = 0
                character_limit = 500000

            return Usage()

        def translate_batch(
            self, texts: list[str], *, source_lang: str, target_lang: str
        ) -> list[str]:
            return [f"KO:{text}" for text in texts]

    monkeypatch.setattr("harmony_translate.pipeline.DeepLClient", FakeDeepLClient)

    config = AppConfig(
        input_path=input_path,
        output_dir=tmp_path / "out",
        sheet_name=None,
        selected_columns=[],
        preserve_original_sheet=False,
        mapped_cell_mode="original_and_translation",
        glossary_path=Path("/Users/usabatch/coding/hazbin_project/glossary.tsv"),
        exclude_patterns_path=Path(
            "/Users/usabatch/coding/hazbin_project/exclude_patterns.yaml"
        ),
        target_lang="KO",
        source_lang="EN",
        provider="deepl",
        deepl_api_key="test-key",
        deepl_base_url="https://api-free.deepl.com",
        gemini_api_key="",
        gemini_model="gemini-2.0-flash",
        gemini_base_url="https://generativelanguage.googleapis.com",
        cache_path=tmp_path / "cache.sqlite3",
    )

    result = run_pipeline(config)
    mapped = load_workbook(result.source_mapped_path)
    mapped_sheet = mapped["Sheet1"]

    assert mapped_sheet["P14"].value == (
        "No character Anim. Add eases to camera move.\n---\n"
        "KO:No character Anim. Add eases to camera move."
    )


def test_apply_mapped_cell_value_keeps_original_style_and_colors_translation_blue() -> (
    None
):
    from harmony_translate.pipeline import _apply_mapped_cell_value

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    cell = sheet["A1"]
    cell.value = "Original"

    _apply_mapped_cell_value(
        cell,
        original_text="Original",
        translated_text="Translated",
        mapped_cell_mode="original_and_translation",
    )

    assert isinstance(cell.value, CellRichText)
    rich_text = cell.value
    assert str(rich_text) == "Original\n---\nTranslated"
    translated_block = rich_text[-1]
    assert isinstance(translated_block, TextBlock)
    assert translated_block.font.color is not None
    assert translated_block.font.color.rgb == "FF0000FF"


def test_pipeline_supports_gemini_provider(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    workbook = build_sample_workbook()
    input_path = tmp_path / "sample.xlsx"
    workbook.save(input_path)

    class FakeGeminiClient:
        def __init__(self, api_key: str, model: str, base_url: str) -> None:
            self.api_key = api_key
            self.model = model
            self.base_url = base_url

        def usage(self):
            return None

        def translate_batch(
            self, texts: list[str], *, source_lang: str, target_lang: str
        ) -> list[str]:
            return [f"GEMINI:{text}" for text in texts]

    monkeypatch.setattr("harmony_translate.pipeline.GeminiClient", FakeGeminiClient)

    config = AppConfig(
        input_path=input_path,
        output_dir=tmp_path / "out",
        sheet_name=None,
        selected_columns=[],
        preserve_original_sheet=False,
        mapped_cell_mode="translation_only",
        glossary_path=Path("/Users/usabatch/coding/hazbin_project/glossary.tsv"),
        exclude_patterns_path=Path(
            "/Users/usabatch/coding/hazbin_project/exclude_patterns.yaml"
        ),
        target_lang="KO",
        source_lang="EN",
        provider="gemini",
        deepl_api_key="",
        deepl_base_url="https://api-free.deepl.com",
        gemini_api_key="gemini-key",
        gemini_model="gemini-2.0-flash",
        gemini_base_url="https://generativelanguage.googleapis.com",
        cache_path=tmp_path / "cache.sqlite3",
    )

    result = run_pipeline(config)
    usage_report = result.usage_path.read_text(encoding="utf-8")

    assert result.preview_mode is False
    assert result.translated_path.exists()
    assert '"provider": "gemini"' in usage_report


def test_pipeline_emits_realtime_logs(tmp_path: Path) -> None:
    workbook = build_sample_workbook()
    input_path = tmp_path / "sample.xlsx"
    workbook.save(input_path)

    config = AppConfig(
        input_path=input_path,
        output_dir=tmp_path / "out",
        sheet_name=None,
        selected_columns=[],
        preserve_original_sheet=False,
        mapped_cell_mode="translation_only",
        glossary_path=Path("/Users/usabatch/coding/hazbin_project/glossary.tsv"),
        exclude_patterns_path=Path(
            "/Users/usabatch/coding/hazbin_project/exclude_patterns.yaml"
        ),
        target_lang="KO",
        source_lang="EN",
        provider="deepl",
        deepl_api_key="",
        deepl_base_url="https://api-free.deepl.com",
        gemini_api_key="",
        gemini_model="gemini-3-flash",
        gemini_base_url="https://generativelanguage.googleapis.com",
        cache_path=tmp_path / "cache.sqlite3",
    )

    logs: list[str] = []
    run_pipeline(config, log_callback=logs.append)

    assert any("Pipeline started" in message for message in logs)
    assert any("Selected columns" in message for message in logs)
    assert any("Saved usage report" in message for message in logs)


def test_pipeline_emits_progress_updates(tmp_path: Path) -> None:
    workbook = build_sample_workbook()
    input_path = tmp_path / "sample.xlsx"
    workbook.save(input_path)

    config = AppConfig(
        input_path=input_path,
        output_dir=tmp_path / "out",
        sheet_name=None,
        selected_columns=[],
        preserve_original_sheet=False,
        mapped_cell_mode="translation_only",
        glossary_path=Path("/Users/usabatch/coding/hazbin_project/glossary.tsv"),
        exclude_patterns_path=Path(
            "/Users/usabatch/coding/hazbin_project/exclude_patterns.yaml"
        ),
        target_lang="KO",
        source_lang="EN",
        provider="deepl",
        deepl_api_key="",
        deepl_base_url="https://api-free.deepl.com",
        gemini_api_key="",
        gemini_model="gemini-3-flash",
        gemini_base_url="https://generativelanguage.googleapis.com",
        cache_path=tmp_path / "cache.sqlite3",
    )

    progress_values: list[int] = []
    run_pipeline(config, progress_callback=progress_values.append)

    assert progress_values[0] == 0
    assert progress_values[-1] == 100
    assert progress_values == sorted(progress_values)
