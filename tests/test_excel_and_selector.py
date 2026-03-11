from openpyxl import Workbook

from harmony_translate.column_selector import (
    profile_columns,
    select_translation_columns,
)
from harmony_translate.excel_io import (
    build_column_label,
    build_sheet_context,
    detect_header_row,
)
from harmony_translate.pipeline import _resolve_selected_profiles


def build_sample_workbook() -> Workbook:
    workbook = Workbook()
    ws = workbook.active
    assert ws is not None
    ws.title = "Sheet1"
    ws["U1"] = "Rough Animation Breakdown"
    ws["A11"] = "Total Count"
    ws["A13"] = "SHOT CODE"
    ws["P13"] = "ANIMATION"
    ws["S13"] = "COMP - OS NOTE"
    ws["Y13"] = "NOTES"
    ws["AN13"] = "COMP TIER"
    ws["A14"] = "HH0304_010_0010"
    ws["P14"] = "No character Anim. Add eases to camera move."
    ws["S14"] = "CHECK:\n- Check Camera Move & add Easing"
    ws["Y14"] = "No character Anim. Add eases to camera move."
    ws["AN14"] = "B"
    ws["A15"] = "HH0304_010_0020"
    ws["P15"] = "Natural acting note for rough animation."
    ws["S15"] = "CHECK:\n- Add blur on rimlights"
    ws["Y15"] = "Natural acting note for rough animation."
    ws["AN15"] = "C"
    ws["A16"] = "HH0304_010_0030"
    ws["P16"] = "Pickup this section with limited animation."
    ws["S16"] = "SUP NOTE:\n- Keep comp subtle"
    ws["Y16"] = "Pickup this section with limited animation."
    ws["AN16"] = "A"
    ws["A17"] = "HH0304_010_0040"
    ws["P17"] = "Storybook style shot, mostly tweens."
    ws["S17"] = "COMP TIER:\n- Standard compositing needed"
    ws["Y17"] = "Storybook style shot, mostly tweens."
    ws["AN17"] = "B"
    ws["A18"] = "HH0304_010_0050"
    ws["P18"] = "Hold the pose longer before the turn."
    ws["S18"] = "USE ASSETS LINKED:\n- CP-VtINTValBedroomDark"
    ws["Y18"] = "Hold the pose longer before the turn."
    ws["AN18"] = "C"
    return workbook


def test_detect_header_row_prefers_shot_code() -> None:
    workbook = build_sample_workbook()
    worksheet = workbook.active
    assert worksheet is not None
    assert detect_header_row(worksheet) == 13


def test_selector_prefers_note_like_columns_and_excludes_tier() -> None:
    workbook = build_sample_workbook()
    context = build_sheet_context(workbook, None)
    profiles = profile_columns(
        context.worksheet,
        header_row=context.header_row,
        data_start_row=context.data_start_row,
    )
    selected = select_translation_columns(profiles)
    selected_headers = {profile.header for profile in selected}
    assert "ANIMATION" in selected_headers
    assert "COMP - OS NOTE" in selected_headers
    assert "NOTES" in selected_headers
    assert "COMP TIER" not in selected_headers


def test_selector_excludes_asset_list_column() -> None:
    workbook = build_sample_workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet["C13"] = "ALL TAGGED ASSETS"
    worksheet["C14"] = "BG-VtINTValBedroomBedESTKey, CP-VtINTValBedroomDark"
    worksheet["C15"] = "BG-VtINTValBedroomBedESTKey, CP-VtINTValBedroomDark"
    worksheet["C16"] = "BG-VtINTValBedroomBedESTKey, CP-VtINTValBedroomDark"
    worksheet["C17"] = "BG-VtINTValBedroomBedESTKey, CP-VtINTValBedroomDark"
    worksheet["C18"] = "BG-VtINTValBedroomBedESTKey, CP-VtINTValBedroomDark"

    context = build_sheet_context(workbook, None)
    profiles = profile_columns(
        context.worksheet,
        header_row=context.header_row,
        data_start_row=context.data_start_row,
    )
    selected = select_translation_columns(profiles)
    selected_headers = {profile.header for profile in selected}
    assert "ALL TAGGED ASSETS" not in selected_headers


def test_pipeline_resolves_column_selection_by_unique_column_label() -> None:
    workbook = build_sample_workbook()
    context = build_sheet_context(workbook, None)
    profiles = profile_columns(
        context.worksheet,
        header_row=context.header_row,
        data_start_row=context.data_start_row,
    )
    auto_selected = select_translation_columns(profiles)
    target_profile = next(profile for profile in profiles if profile.index == 25)
    selected = _resolve_selected_profiles(
        auto_selected,
        profiles,
        [build_column_label(target_profile.index, target_profile.header)],
    )
    assert [profile.index for profile in selected] == [25]
