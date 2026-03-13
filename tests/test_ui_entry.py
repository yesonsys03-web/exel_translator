from pathlib import Path
import importlib.util
import os

import pytest
from openpyxl import Workbook
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QTableWidgetItem

from harmony_translate.cli import DEFAULT_INPUT_PATH, build_parser
from harmony_translate.glossary import load_glossary
from harmony_translate.ui import GLOSSARY_CANDIDATES_ROLE, MainWindow, USER_ROLE


def load_main_module():
    module_path = Path(__file__).resolve().parents[1] / "main.py"
    spec = importlib.util.spec_from_file_location("main_module", module_path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def test_parser_default_input_still_matches_sample() -> None:
    parser = build_parser()
    args = parser.parse_args([])
    assert args.input == str(DEFAULT_INPUT_PATH)
    assert args.provider == "gemini"
    assert args.gemini_model == "gemini-2.5-flash"


def test_ui_does_not_offer_deepl_provider(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Alpha"
    path = tmp_path / "ui-provider.xlsx"
    workbook.save(path)

    window = MainWindow()
    window.input_edit.setText(str(path))
    window.load_workbook_preview()

    provider_values = [
        window.provider_combo.itemData(index)
        for index in range(window.provider_combo.count())
    ]
    assert provider_values == ["gemini"]

    window.close()
    app.quit()


def test_ui_offers_deepl_provider_when_feature_flag_enabled(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Alpha"
    path = tmp_path / "ui-provider-enabled.xlsx"
    workbook.save(path)

    monkeypatch.setenv("TRANSLATION_ENABLE_DEEPL", "true")

    window = MainWindow()
    window.input_edit.setText(str(path))
    window.load_workbook_preview()

    provider_values = [
        window.provider_combo.itemData(index)
        for index in range(window.provider_combo.count())
    ]
    assert provider_values == ["gemini", "deepl"]

    window.close()
    app.quit()


def test_main_exposes_gui_launcher() -> None:
    module = load_main_module()
    assert callable(module.launch_ui)


def test_main_window_populates_sheet_selector_from_workbook(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    workbook.create_sheet("Notes")
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Test note"
    path = tmp_path / "ui-sheets.xlsx"
    workbook.save(path)

    window = MainWindow()
    window.input_edit.setText(str(path))
    window.load_workbook_preview()

    sheet_names = [
        window.sheet_combo.itemText(index)
        for index in range(window.sheet_combo.count())
    ]
    assert sheet_names == ["Sheet1", "Notes"]
    assert window.get_selected_sheet_name() == "Sheet1"
    column_labels = []
    for index in range(window.column_list.count()):
        item = window.column_list.item(index)
        assert item is not None
        column_labels.append(item.text())
    assert any(label.startswith("A | SHOT CODE") for label in column_labels)
    assert any(label.startswith("P | ANIMATION") for label in column_labels)
    assert window.preview_table.columnCount() >= 2
    assert window.preview_table.rowCount() >= 1
    header_labels = []
    for index in range(window.preview_table.columnCount()):
        header_item = window.preview_table.horizontalHeaderItem(index)
        if header_item is None:
            continue
        header_labels.append(header_item.text())
    assert any(label.endswith("SHOT CODE") for label in header_labels)
    assert any(label.endswith("ANIMATION") for label in header_labels)
    first_row_value = window.preview_table.item(0, 0)
    assert first_row_value is not None
    assert first_row_value.text() == "HH0304_010_0010"
    window.close()
    app.quit()


def test_column_list_selection_focuses_matching_preview_column(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["Y13"] = "NOTES"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Test animation note"
    sheet1["Y14"] = "Test notes column"
    path = tmp_path / "ui-focus.xlsx"
    workbook.save(path)

    window = MainWindow()
    window.input_edit.setText(str(path))
    window.load_workbook_preview()

    target_item = None
    for index in range(window.column_list.count()):
        item = window.column_list.item(index)
        assert item is not None
        if str(item.data(USER_ROLE)).startswith("P | ANIMATION"):
            target_item = item
            break
    assert target_item is not None

    window.column_list.setCurrentItem(target_item)

    current_column = window.preview_table.currentColumn()
    header_item = window.preview_table.horizontalHeaderItem(current_column)
    assert header_item is not None
    assert header_item.text().startswith("P | ANIMATION")
    window.close()
    app.quit()


def test_selected_column_character_count_updates_live(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["Y13"] = "NOTES"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Alpha"
    sheet1["Y14"] = "BetaBeta"
    sheet1["A15"] = "HH0304_010_0020"
    sheet1["P15"] = "Gamma"
    sheet1["Y15"] = "Delta"
    path = tmp_path / "ui-char-count.xlsx"
    workbook.save(path)

    window = MainWindow()
    window.input_edit.setText(str(path))
    window.load_workbook_preview()

    target_item = None
    for index in range(window.column_list.count()):
        item = window.column_list.item(index)
        assert item is not None
        if str(item.data(USER_ROLE)).startswith("P | ANIMATION"):
            target_item = item
            break
    assert target_item is not None

    target_item.setCheckState(Qt.CheckState.Unchecked)
    text_without_p = window.selection_stats_label.text()
    target_item.setCheckState(Qt.CheckState.Checked)
    text_with_p = window.selection_stats_label.text()

    assert "총 글자 수" in text_without_p
    assert "총 글자 수" in text_with_p
    assert text_with_p != text_without_p

    window.close()
    app.quit()


def test_deepl_env_provider_falls_back_to_gemini_in_ui(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Alpha"
    path = tmp_path / "ui-dynamic-limit.xlsx"
    workbook.save(path)

    monkeypatch.setenv("TRANSLATION_PROVIDER", "deepl")

    window = MainWindow()
    window.input_edit.setText(str(path))
    window.load_workbook_preview()

    assert window._current_provider() == "gemini"
    assert window.provider_combo.currentData() == "gemini"

    window.close()
    app.quit()


def test_gemini_provider_shows_model_token_limits(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Alpha"
    path = tmp_path / "ui-gemini-limit.xlsx"
    workbook.save(path)

    class FakeModel:
        def __init__(
            self, model_id: str, display_name: str, input_limit: int, output_limit: int
        ) -> None:
            self.model_id = model_id
            self.display_name = display_name
            self.input_token_limit = input_limit
            self.output_token_limit = output_limit

    class FakeGeminiClient:
        def __init__(self, api_key: str, model: str, base_url: str) -> None:
            self.api_key = api_key
            self.model = model
            self.base_url = base_url

        def list_models(self):
            return [
                FakeModel("gemini-2.0-flash", "Gemini 2.0 Flash", 1048576, 8192),
                FakeModel(
                    "gemini-2.0-flash-lite", "Gemini 2.0 Flash Lite", 1048576, 8192
                ),
            ]

    monkeypatch.setenv("GEMINI_API_KEY", "gemini-key")
    monkeypatch.setattr("harmony_translate.ui.GeminiClient", FakeGeminiClient)

    window = MainWindow()
    window.input_edit.setText(str(path))
    provider_index = window.provider_combo.findData("gemini")
    assert provider_index >= 0
    window.provider_combo.setCurrentIndex(provider_index)
    window.load_workbook_preview()

    stats_text = window.selection_stats_label.text()
    assert "Gemini gemini-2.0-flash 제한 입력 1,048,576tok" in stats_text
    assert "출력 8,192tok" in stats_text

    window.close()
    app.quit()


def test_gemini_model_candidates_env_controls_fallback_model_list(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Alpha"
    path = tmp_path / "ui-gemini-model-candidates.xlsx"
    workbook.save(path)

    class FakeGeminiClient:
        def __init__(self, api_key: str, model: str, base_url: str) -> None:
            self.api_key = api_key
            self.model = model
            self.base_url = base_url

        def list_models(self):
            return []

    monkeypatch.setenv("GEMINI_API_KEY", "gemini-key")
    monkeypatch.setenv(
        "GEMINI_MODEL_CANDIDATES",
        "gemini-2.5-pro, gemini-3.1-pro-preview, gemini-2.5-flash",
    )
    monkeypatch.setattr("harmony_translate.ui.GeminiClient", FakeGeminiClient)

    window = MainWindow()
    window.input_edit.setText(str(path))
    provider_index = window.provider_combo.findData("gemini")
    assert provider_index >= 0
    window.provider_combo.setCurrentIndex(provider_index)
    window.load_workbook_preview()

    model_values = [
        window.model_combo.itemData(index)
        for index in range(window.model_combo.count())
    ]
    assert "gemini-2.5-pro" in model_values
    assert "gemini-3.1-pro-preview" in model_values
    assert "gemini-2.5-flash" in model_values

    window.close()
    app.quit()


def test_gemini_dropdown_hides_legacy_models_when_primary_models_available(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Alpha"
    path = tmp_path / "ui-gemini-curated-models.xlsx"
    workbook.save(path)

    class FakeModel:
        def __init__(
            self, model_id: str, display_name: str, input_limit: int, output_limit: int
        ) -> None:
            self.model_id = model_id
            self.display_name = display_name
            self.input_token_limit = input_limit
            self.output_token_limit = output_limit

    class FakeGeminiClient:
        def __init__(self, api_key: str, model: str, base_url: str) -> None:
            self.api_key = api_key
            self.model = model
            self.base_url = base_url

        def list_models(self):
            return [
                FakeModel("gemini-2.0-flash", "Gemini 2.0 Flash", 1048576, 8192),
                FakeModel("gemini-2.5-flash", "Gemini 2.5 Flash", 1048576, 65536),
                FakeModel("gemini-2.5-pro", "Gemini 2.5 Pro", 1048576, 65536),
                FakeModel(
                    "gemini-3.1-pro-preview",
                    "Gemini 3.1 Pro Preview",
                    1048576,
                    65536,
                ),
            ]

    monkeypatch.setenv("GEMINI_API_KEY", "gemini-key")
    monkeypatch.setattr("harmony_translate.ui.GeminiClient", FakeGeminiClient)

    window = MainWindow()
    window.input_edit.setText(str(path))
    provider_index = window.provider_combo.findData("gemini")
    assert provider_index >= 0
    window.provider_combo.setCurrentIndex(provider_index)
    window.load_workbook_preview()

    model_values = [
        window.model_combo.itemData(index)
        for index in range(window.model_combo.count())
    ]
    assert "gemini-2.5-flash" in model_values
    assert "gemini-2.5-flash-lite" in model_values
    assert "gemini-2.5-pro" in model_values
    assert "gemini-3.1-pro-preview" in model_values
    assert "gemini-2.0-flash" not in model_values

    window.close()
    app.quit()


def test_progress_bar_updates_from_worker_signal(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Alpha"
    path = tmp_path / "ui-progress.xlsx"
    workbook.save(path)

    window = MainWindow()
    window.input_edit.setText(str(path))
    window.load_workbook_preview()

    window.handle_run_progress_value(42)

    assert window.progress_bar.value() == 42
    assert window.progress_bar.format() == "진행률 42%"

    window.close()
    app.quit()


def test_preserve_original_checkbox_defaults_enabled(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Alpha"
    path = tmp_path / "ui-preserve-option.xlsx"
    workbook.save(path)

    window = MainWindow()
    window.input_edit.setText(str(path))
    window.load_workbook_preview()

    assert window.preserve_original_checkbox.isChecked() is True
    assert "Sheet1_ORIGINAL" in window.preserve_original_checkbox.text()

    workbook2 = Workbook()
    second_sheet = workbook2.active
    assert second_sheet is not None
    second_sheet.title = "Dialogue"
    second_sheet["A13"] = "SHOT CODE"
    second_sheet["P13"] = "ANIMATION"
    second_sheet["A14"] = "HH0304_010_0010"
    second_sheet["P14"] = "Beta"
    path2 = tmp_path / "ui-preserve-option-2.xlsx"
    workbook2.save(path2)

    window.input_edit.setText(str(path2))
    window.load_workbook_preview()

    assert "Dialogue_ORIGINAL" in window.preserve_original_checkbox.text()
    assert window.mapped_cell_mode_combo.currentData() == "translation_only"

    window.close()
    app.quit()


def test_ui_infers_project_id_from_input_filename(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Alpha"
    path = tmp_path / "HH0304-sample.xlsx"
    workbook.save(path)

    window = MainWindow()
    window.project_id_edit.setText("")
    window.input_edit.setText(str(path))
    window.load_workbook_preview()

    assert window.project_id_edit.text() == "HH0304"

    window.close()
    app.quit()


def test_ui_saves_project_glossary_to_project_path(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Alpha"
    path = tmp_path / "HH0304-ui-glossary.xlsx"
    workbook.save(path)

    window = MainWindow()
    window.input_edit.setText(str(path))
    window.load_workbook_preview()
    window.project_id_edit.setText("TESTPROJ")
    window.add_glossary_row()

    source_item = window.glossary_table.item(0, 0)
    target_item = window.glossary_table.item(0, 1)
    assert source_item is not None
    assert target_item is not None
    source_item.setText("Node View")
    target_item.setText("노드 뷰")

    previous_cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        window.save_glossary_editor()
        project_glossary_path = (
            tmp_path / "glossary" / "projects" / "TESTPROJ" / "glossary.tsv"
        )
        assert project_glossary_path.exists()
        glossary = load_glossary(project_glossary_path)
        assert glossary == {"Node View": "노드 뷰"}
    finally:
        os.chdir(previous_cwd)

    window.close()
    app.quit()


def test_ui_auto_populates_glossary_candidates_from_source(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["Y13"] = "NOTES"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Open Node View and check SFX timing."
    sheet1["Y14"] = "Composite pass in Node View."
    path = tmp_path / "auto-candidate.xlsx"
    workbook.save(path)

    window = MainWindow()
    window.project_id_edit.setText("AUTOCASE")

    previous_cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        window.input_edit.setText(str(path))
        window.load_workbook_preview()
    finally:
        os.chdir(previous_cwd)

    assert window.glossary_table.rowCount() > 0
    candidates: set[str] = set()
    for index in range(window.glossary_table.rowCount()):
        item = window.glossary_table.item(index, 0)
        if item is None:
            continue
        candidates.add(item.text())
    assert "Node View" in candidates or "SFX" in candidates

    window.close()
    app.quit()


def test_ui_generate_button_populates_candidates_for_lowercase_text(
    tmp_path: Path,
) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "camera movement needs smoother acting"
    sheet1["A15"] = "HH0304_010_0020"
    sheet1["P15"] = "acting timing should match camera movement"
    path = tmp_path / "manual-generate.xlsx"
    workbook.save(path)

    window = MainWindow()
    previous_cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        window.input_edit.setText(str(path))
        window.project_id_edit.setText("MANUALCASE")
        window.load_workbook_preview()
        window.glossary_table.setRowCount(0)
        window.generate_glossary_candidates()
    finally:
        os.chdir(previous_cwd)

    assert window.glossary_table.rowCount() > 0

    window.close()
    app.quit()


def test_glossary_cell_click_creates_missing_item(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Alpha"
    path = tmp_path / "click-edit.xlsx"
    workbook.save(path)

    window = MainWindow()
    window.input_edit.setText(str(path))
    window.load_workbook_preview()
    window.glossary_table.setRowCount(1)
    window.glossary_table.setItem(0, 0, None)
    window.glossary_table.setItem(0, 1, None)

    window.handle_glossary_cell_clicked(0, 1)

    item = window.glossary_table.item(0, 1)
    assert item is not None

    window.close()
    app.quit()


def test_glossary_table_disables_elide_and_enables_wrap() -> None:
    app = QApplication.instance() or QApplication([])
    window = MainWindow()

    assert window.glossary_table.wordWrap() is True
    assert window.glossary_table.textElideMode() == Qt.TextElideMode.ElideNone

    window.close()
    app.quit()


def test_generate_candidates_prefills_translation_from_global_glossary(
    tmp_path: Path,
) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Open Node View for final composite."
    path = tmp_path / "prefill-glossary.xlsx"
    workbook.save(path)

    glossary_root = tmp_path / "glossary"
    glossary_root.mkdir(parents=True, exist_ok=True)
    (glossary_root / "global.tsv").write_text("Node View\t노드 뷰\n", encoding="utf-8")

    window = MainWindow()
    previous_cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        window.input_edit.setText(str(path))
        window.project_id_edit.setText("PREFILLCASE")
        window.load_workbook_preview()
        window.glossary_table.setRowCount(0)
        window.generate_glossary_candidates()
    finally:
        os.chdir(previous_cwd)

    matched_translation = ""
    for index in range(window.glossary_table.rowCount()):
        source_item = window.glossary_table.item(index, 0)
        target_item = window.glossary_table.item(index, 1)
        if source_item is None or target_item is None:
            continue
        if source_item.text() == "Node View":
            matched_translation = target_item.text()
            break

    assert matched_translation == "노드 뷰"

    window.close()
    app.quit()


def test_glossary_prefill_does_not_use_llm_by_default(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "camera movement needs smoother acting"
    path = tmp_path / "no-llm-prefill.xlsx"
    workbook.save(path)

    monkeypatch.delenv("GLOSSARY_PREFILL_ALLOW_LLM", raising=False)
    monkeypatch.setenv("GEMINI_API_KEY", "dummy")

    def fail_translate(*args, **kwargs):
        raise AssertionError("LLM prefill should be disabled by default")

    monkeypatch.setattr(
        "harmony_translate.ui.GeminiClient.translate_batch",
        fail_translate,
    )

    window = MainWindow()
    previous_cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        window.input_edit.setText(str(path))
        window.project_id_edit.setText("NO_LLM")
        window.load_workbook_preview()
        window.glossary_table.setRowCount(0)
        window.generate_glossary_candidates()
    finally:
        os.chdir(previous_cwd)

    assert window.glossary_table.rowCount() > 0
    non_empty_translation_found = False
    for index in range(window.glossary_table.rowCount()):
        target_item = window.glossary_table.item(index, 1)
        if target_item is None:
            continue
        if target_item.text().strip():
            non_empty_translation_found = True
            break
    assert non_empty_translation_found

    window.close()
    app.quit()


def test_candidate_options_stored_for_translation_cell(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Node View setup and camera movement"
    path = tmp_path / "candidate-options.xlsx"
    workbook.save(path)

    window = MainWindow()
    previous_cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        window.input_edit.setText(str(path))
        window.project_id_edit.setText("OPTCASE")
        window.generate_glossary_candidates()
    finally:
        os.chdir(previous_cwd)

    assert window.glossary_table.rowCount() > 0
    assert window.glossary_table.columnCount() == 3
    target_item = window.glossary_table.item(0, 1)
    assert target_item is not None
    raw_options = target_item.data(GLOSSARY_CANDIDATES_ROLE)
    assert isinstance(raw_options, list)
    assert len(raw_options) > 0
    assert window.glossary_table.cellWidget(0, 2) is not None

    window.close()
    app.quit()


def test_candidate_options_have_multiple_choices_without_llm(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIMATION"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Open Node View for camera movement timing"
    path = tmp_path / "candidate-multi-options.xlsx"
    workbook.save(path)

    window = MainWindow()
    previous_cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        window.input_edit.setText(str(path))
        window.project_id_edit.setText("MULTIOPT")
        window.generate_glossary_candidates()
    finally:
        os.chdir(previous_cwd)

    assert window.glossary_table.rowCount() > 0
    found_multi_option = False
    for index in range(window.glossary_table.rowCount()):
        target_item = window.glossary_table.item(index, 1)
        if target_item is None:
            continue
        raw_options = target_item.data(GLOSSARY_CANDIDATES_ROLE)
        if isinstance(raw_options, list) and len(raw_options) >= 2:
            found_multi_option = True
            break
    assert found_multi_option

    window.close()
    app.quit()


def test_generate_candidates_uses_checked_columns_only(tmp_path: Path) -> None:
    app = QApplication.instance() or QApplication([])
    workbook = Workbook()
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1["A13"] = "SHOT CODE"
    sheet1["P13"] = "ANIM_A"
    sheet1["Y13"] = "ANIM_B"
    sheet1["A14"] = "HH0304_010_0010"
    sheet1["P14"] = "Node View pass"
    sheet1["Y14"] = "Camera movement timing"
    path = tmp_path / "checked-columns-only.xlsx"
    workbook.save(path)

    window = MainWindow()
    previous_cwd = Path.cwd()
    try:
        os.chdir(tmp_path)
        window.input_edit.setText(str(path))
        window.load_workbook_preview()
        for index in range(window.column_list.count()):
            item = window.column_list.item(index)
            if item is None:
                continue
            label = str(item.data(USER_ROLE) or "")
            if "ANIM_A" in label:
                item.setCheckState(Qt.CheckState.Checked)
            elif "ANIM_B" in label:
                item.setCheckState(Qt.CheckState.Unchecked)
        window.glossary_table.setRowCount(0)
        window.generate_glossary_candidates()
    finally:
        os.chdir(previous_cwd)

    candidates: set[str] = set()
    for index in range(window.glossary_table.rowCount()):
        source_item = window.glossary_table.item(index, 0)
        if source_item is None:
            continue
        candidates.add(source_item.text())

    assert "Node View" in candidates
    assert "Camera" not in candidates

    window.close()
    app.quit()


def test_click_translation_cell_uses_candidate_picker(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    app = QApplication.instance() or QApplication([])
    window = MainWindow()
    window.glossary_table.setRowCount(1)
    source_item = QTableWidgetItem("Node View")
    target_item = QTableWidgetItem("노드 뷰")
    target_item.setData(GLOSSARY_CANDIDATES_ROLE, ["노드 뷰", "노드 화면", "노드 창"])
    window.glossary_table.setItem(0, 0, source_item)
    window.glossary_table.setItem(0, 1, target_item)
    window._set_candidate_button(0, ["노드 뷰", "노드 화면", "노드 창"])

    monkeypatch.setattr(
        "harmony_translate.ui.QInputDialog.getItem",
        lambda *args, **kwargs: ("노드 화면", True),
    )

    window._show_translation_candidates(0)

    result_item = window.glossary_table.item(0, 1)
    assert result_item is not None
    assert result_item.text() == "노드 화면"

    window.close()
    app.quit()
