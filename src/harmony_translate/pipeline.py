from __future__ import annotations

from dataclasses import dataclass
from copy import copy
from pathlib import Path
import re
from typing import Any, Callable

from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Color, Font
import yaml

from harmony_translate.audit import AuditEntry, export_audit, export_usage_report
from harmony_translate.cache import TranslationCache
from harmony_translate.column_selector import (
    ColumnProfile,
    profile_columns,
    select_translation_columns,
)
from harmony_translate.config import AppConfig
from harmony_translate.excel_io import (
    SheetContext,
    append_translation_columns,
    build_column_label,
    build_sheet_context,
    load_excel_workbook,
    save_workbook,
)
from harmony_translate.glossary import apply_term_locks, load_glossary
from harmony_translate.preprocess import (
    build_deduplicated_texts,
    looks_like_code,
    normalize_text,
)
from harmony_translate.translator_deepl import DeepLClient
from harmony_translate.translator_gemini import GeminiClient


@dataclass
class PipelineResult:
    translated_path: Path
    source_mapped_path: Path
    audit_path: Path
    usage_path: Path
    selected_headers: list[str]
    preview_mode: bool


LogCallback = Callable[[str], None]
ProgressCallback = Callable[[int], None]
TRANSLATED_FONT_COLOR = "FF0000FF"


def load_exclude_patterns(path: Path) -> list[re.Pattern[str]]:
    # [ANCHOR:PIPELINE_LOAD_EXCLUDE_PATTERNS]
    if not path.exists():
        return []
    payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    if not isinstance(payload, dict):
        return []
    raw_patterns = payload.get("patterns", [])
    if not isinstance(raw_patterns, list):
        return []
    return [re.compile(str(pattern)) for pattern in raw_patterns]


def run_pipeline(
    config: AppConfig,
    *,
    log_callback: LogCallback | None = None,
    progress_callback: ProgressCallback | None = None,
) -> PipelineResult:
    # [ANCHOR:PIPELINE_RUN]
    _emit_log(log_callback, f"Pipeline started (provider={config.provider})")
    workbook = load_excel_workbook(config.input_path)
    mapped_workbook = load_excel_workbook(config.input_path)
    context = build_sheet_context(workbook, config.sheet_name)
    mapped_sheet_name = config.sheet_name or mapped_workbook.sheetnames[0]
    if config.preserve_original_sheet:
        source_sheet = mapped_workbook[mapped_sheet_name]
        backup_sheet = mapped_workbook.copy_worksheet(source_sheet)
        backup_sheet.title = _build_original_sheet_name(
            mapped_workbook, mapped_sheet_name
        )
    glossary = load_glossary(config.glossary_path)
    exclude_patterns = load_exclude_patterns(config.exclude_patterns_path)
    profiles = profile_columns(
        context.worksheet,
        header_row=context.header_row,
        data_start_row=context.data_start_row,
    )
    auto_selected = select_translation_columns(profiles)
    selected_profiles = _resolve_selected_profiles(
        auto_selected, profiles, config.selected_columns
    )
    selected_column_indexes = [profile.index for profile in selected_profiles]
    selected_headers = [profile.header for profile in selected_profiles]
    _emit_log(log_callback, f"Selected columns: {len(selected_headers)}")
    _emit_progress(progress_callback, 0)
    translation_columns = append_translation_columns(
        context.worksheet, context.header_row, selected_column_indexes
    )
    mapped_context = build_sheet_context(mapped_workbook, config.sheet_name)

    cache = TranslationCache(config.cache_path)
    client = _build_translation_client(config)
    _emit_log(
        log_callback,
        "Preview mode enabled, API calls skipped"
        if client is None
        else "Translation client initialized",
    )
    cache_namespace = _build_cache_namespace(config)
    usage_before_count = 0
    usage_limit = 0
    if client is not None:
        usage_before = client.usage()
        if usage_before is not None:
            usage_before_count = usage_before.character_count
            usage_limit = usage_before.character_limit
    audit_entries: list[AuditEntry] = []

    try:
        total_columns = len(translation_columns)
        completed_columns = 0
        for source_column, target_column in translation_columns.items():
            _emit_log(
                log_callback,
                f"Translating column {_column_label(context, source_column)} -> {_column_label(context, target_column)}",
            )
            _translate_column(
                context=context,
                source_column=source_column,
                target_column=target_column,
                glossary=glossary,
                exclude_patterns=exclude_patterns,
                cache=cache,
                client=client,
                source_lang=config.source_lang,
                target_lang=config.target_lang,
                audit_entries=audit_entries,
                mapped_context=mapped_context,
                mapped_cell_mode=config.mapped_cell_mode,
                cache_namespace=cache_namespace,
                log_callback=log_callback,
            )
            completed_columns += 1
            progress_percent = int((completed_columns / max(total_columns, 1)) * 100)
            _emit_progress(progress_callback, progress_percent)

        translated_path = config.output_dir / "translated.xlsx"
        source_mapped_path = config.output_dir / "source_mapped.xlsx"
        audit_path = config.output_dir / "translation_audit.xlsx"
        usage_path = config.output_dir / "usage_report.json"
        save_workbook(workbook, translated_path)
        save_workbook(mapped_workbook, source_mapped_path)
        _emit_log(log_callback, f"Saved translated workbook: {translated_path}")
        _emit_log(log_callback, f"Saved source mapped workbook: {source_mapped_path}")
        export_audit(audit_entries, audit_path)
        _emit_log(log_callback, f"Saved audit workbook: {audit_path}")
        usage_after_count = usage_before_count
        if client is not None:
            usage_after = client.usage()
            if usage_after is not None:
                usage_after_count = usage_after.character_count
        export_usage_report(
            {
                "provider": config.provider,
                "gemini_model": config.gemini_model,
                "preview_mode": config.preview_mode,
                "character_count_before": usage_before_count,
                "character_limit": usage_limit,
                "character_count_after": usage_after_count,
                "selected_headers": selected_headers,
                "translated_rows": sum(
                    1 for entry in audit_entries if not entry.skipped
                ),
                "skipped_rows": sum(1 for entry in audit_entries if entry.skipped),
            },
            usage_path,
        )
        _emit_log(log_callback, f"Saved usage report: {usage_path}")
        _emit_progress(progress_callback, 100)
    finally:
        cache.close()
        _emit_log(log_callback, "Pipeline finished")

    return PipelineResult(
        translated_path=translated_path,
        source_mapped_path=source_mapped_path,
        audit_path=audit_path,
        usage_path=usage_path,
        selected_headers=selected_headers,
        preview_mode=config.preview_mode,
    )


def _resolve_selected_profiles(
    auto_selected: list[ColumnProfile],
    all_profiles: list[ColumnProfile],
    selected_columns: list[str],
) -> list[ColumnProfile]:
    if not selected_columns:
        return auto_selected
    selected_lookup = {name.upper() for name in selected_columns}
    return [
        profile
        for profile in all_profiles
        if profile.header.upper() in selected_lookup
        or build_column_label(profile.index, profile.header).upper() in selected_lookup
    ]


def _translate_column(
    *,
    context: SheetContext,
    source_column: int,
    target_column: int,
    glossary: dict[str, str],
    exclude_patterns: list[re.Pattern[str]],
    cache: TranslationCache,
    client: DeepLClient | GeminiClient | None,
    source_lang: str,
    target_lang: str,
    audit_entries: list[AuditEntry],
    mapped_context: SheetContext,
    mapped_cell_mode: str,
    cache_namespace: str,
    log_callback: LogCallback | None,
) -> None:
    # [ANCHOR:PIPELINE_TRANSLATE_COLUMN]
    values_by_row: dict[int, str] = {}
    protected_values_by_row: dict[int, str] = {}
    for row_index in range(context.data_start_row, context.worksheet.max_row + 1):
        cell_value = context.worksheet.cell(row_index, source_column).value
        if cell_value in (None, ""):
            continue
        if isinstance(cell_value, (int, float)):
            audit_entries.append(
                AuditEntry(
                    context.worksheet.title,
                    row_index,
                    context.headers[source_column],
                    str(cell_value),
                    "",
                    False,
                    True,
                    "numeric",
                )
            )
            continue
        normalized = normalize_text(str(cell_value))
        if not normalized:
            continue
        if looks_like_code(normalized, exclude_patterns):
            audit_entries.append(
                AuditEntry(
                    context.worksheet.title,
                    row_index,
                    context.headers[source_column],
                    normalized,
                    normalized,
                    False,
                    True,
                    "excluded_pattern",
                )
            )
            continue
        values_by_row[row_index] = normalized
        protected_values_by_row[row_index] = apply_term_locks(normalized, glossary)

    deduped_texts, _ = build_deduplicated_texts(list(protected_values_by_row.values()))
    cache_key_by_text = {text: f"{cache_namespace}|{text}" for text in deduped_texts}
    cached_by_key = cache.get_many(list(cache_key_by_text.values()))
    cached: dict[str, str] = {
        text: cached_by_key[key]
        for text, key in cache_key_by_text.items()
        if key in cached_by_key
    }
    missing = [text for text in deduped_texts if text not in cached]
    _emit_log(
        log_callback,
        f"Column {context.headers[source_column]}: total={len(deduped_texts)} cache_hit={len(cached)} request={len(missing)}",
    )
    translated_pairs: dict[str, str] = {}
    if client is not None and missing:
        translated = client.translate_batch(
            missing, source_lang=source_lang, target_lang=target_lang
        )
        if len(missing) != len(translated):
            raise RuntimeError("DeepL returned a mismatched translation count")
        translated_pairs = {
            missing[index]: translated[index] for index in range(len(missing))
        }
        cache.set_many(
            {
                cache_key_by_text[text]: translated_text
                for text, translated_text in translated_pairs.items()
            }
        )

    all_translations = {**cached, **translated_pairs}

    for row_index, protected in protected_values_by_row.items():
        translated_text = all_translations.get(protected, "")
        if translated_text:
            translated_cell = context.worksheet.cell(
                row=row_index, column=target_column, value=translated_text
            )
            _apply_translated_font_color(translated_cell)
            mapped_cell = mapped_context.worksheet.cell(
                row=row_index,
                column=source_column,
            )
            _apply_mapped_cell_value(
                mapped_cell,
                original_text=values_by_row[row_index],
                translated_text=translated_text,
                mapped_cell_mode=mapped_cell_mode,
            )
        audit_entries.append(
            AuditEntry(
                context.worksheet.title,
                row_index,
                context.headers[source_column],
                values_by_row[row_index],
                translated_text,
                protected in cached,
                client is None,
                "missing_api_key_preview" if client is None else "",
            )
        )


def _emit_log(log_callback: LogCallback | None, message: str) -> None:
    if log_callback is None:
        return
    log_callback(message)


def _emit_progress(progress_callback: ProgressCallback | None, value: int) -> None:
    if progress_callback is None:
        return
    progress_callback(max(0, min(100, value)))


def _column_label(context: SheetContext, column_index: int) -> str:
    header = context.headers.get(column_index)
    if header:
        return header
    fallback_header = context.worksheet.cell(context.header_row, column_index).value
    return build_column_label(column_index, str(fallback_header or "translation"))


def _apply_translated_font_color(cell) -> None:
    font = copy(cell.font) if cell.font is not None else Font()
    font.color = TRANSLATED_FONT_COLOR
    cell.font = font


def _apply_mapped_cell_value(
    cell,
    *,
    original_text: str,
    translated_text: str,
    mapped_cell_mode: str,
) -> None:
    if mapped_cell_mode == "original_and_translation":
        cell.value = CellRichText(
            original_text,
            "\n---\n",
            TextBlock(_build_translated_inline_font(cell.font), translated_text),
        )
        return

    cell.value = translated_text
    _apply_translated_font_color(cell)


def _build_translated_inline_font(font: Font | None) -> InlineFont:
    base_font = font or Font()
    inline_font = InlineFont(
        rFont=base_font.name,
        charset=base_font.charset,
        family=base_font.family,
        b=base_font.bold,
        i=base_font.italic,
        strike=base_font.strike,
        outline=base_font.outline,
        shadow=base_font.shadow,
        condense=base_font.condense,
        extend=base_font.extend,
        sz=base_font.sz,
        u=base_font.underline,
        vertAlign=base_font.vertAlign,
        scheme=base_font.scheme,
        color=Color(rgb=TRANSLATED_FONT_COLOR),
    )
    return inline_font


def _build_original_sheet_name(workbook, base_name: str) -> str:
    candidate = f"{base_name}_ORIGINAL"
    if candidate not in workbook.sheetnames:
        return candidate
    suffix = 2
    while f"{candidate}_{suffix}" in workbook.sheetnames:
        suffix += 1
    return f"{candidate}_{suffix}"


def _build_translation_client(config: AppConfig):
    if config.preview_mode:
        return None
    provider = config.provider.strip().lower()
    if provider == "gemini":
        return GeminiClient(
            api_key=config.gemini_api_key,
            model=config.gemini_model,
            base_url=config.gemini_base_url,
        )
    return DeepLClient(api_key=config.deepl_api_key, base_url=config.deepl_base_url)


def _build_cache_namespace(config: AppConfig) -> str:
    provider = config.provider.strip().lower()
    if provider == "gemini":
        return f"gemini:{config.gemini_model}:{config.source_lang}:{config.target_lang}"
    return f"deepl:{config.source_lang}:{config.target_lang}"
