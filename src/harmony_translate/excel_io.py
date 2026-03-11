from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from harmony_translate.preprocess import normalize_text


@dataclass
class SheetContext:
    worksheet: Worksheet
    header_row: int
    data_start_row: int
    headers: dict[int, str]


def build_column_label(column_index: int, header: str) -> str:
    # [ANCHOR:EXCEL_BUILD_COLUMN_LABEL]
    return f"{get_column_letter(column_index)} | {header}"


def load_excel_workbook(path: Path) -> Workbook:
    # [ANCHOR:EXCEL_LOAD_WORKBOOK]
    return load_workbook(path)


def detect_header_row(worksheet: Worksheet) -> int:
    # [ANCHOR:EXCEL_DETECT_HEADER_ROW]
    for row_index in range(1, min(worksheet.max_row, 40) + 1):
        row_values = [
            normalize_text(str(worksheet.cell(row_index, col).value or ""))
            for col in range(1, worksheet.max_column + 1)
        ]
        if any(value == "SHOT CODE" for value in row_values):
            return row_index

    best_row = 1
    best_non_empty = -1
    for row_index in range(1, min(worksheet.max_row, 40) + 1):
        non_empty = sum(
            1
            for col in range(1, worksheet.max_column + 1)
            if worksheet.cell(row_index, col).value not in (None, "")
        )
        if non_empty > best_non_empty:
            best_non_empty = non_empty
            best_row = row_index
    return best_row


def build_sheet_context(workbook: Workbook, sheet_name: str | None) -> SheetContext:
    # [ANCHOR:EXCEL_BUILD_SHEET_CONTEXT]
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    header_row = detect_header_row(worksheet)
    data_start_row = header_row + 1
    headers: dict[int, str] = {}
    for column_index in range(1, worksheet.max_column + 1):
        value = worksheet.cell(header_row, column_index).value
        headers[column_index] = normalize_text(str(value or f"COL_{column_index}"))
    return SheetContext(
        worksheet=worksheet,
        header_row=header_row,
        data_start_row=data_start_row,
        headers=headers,
    )


def append_translation_columns(
    worksheet: Worksheet, header_row: int, selected_columns: list[int]
) -> dict[int, int]:
    # [ANCHOR:EXCEL_APPEND_TRANSLATION_COLUMNS]
    mapping: dict[int, int] = {}
    start_column = worksheet.max_column + 1
    for offset, source_column in enumerate(selected_columns):
        target_column = start_column + offset
        source_header = normalize_text(
            str(
                worksheet.cell(header_row, source_column).value
                or f"COL_{source_column}"
            )
        )
        worksheet.cell(
            row=header_row, column=target_column, value=f"{source_header}_KR"
        )
        mapping[source_column] = target_column
    return mapping


def save_workbook(workbook: Workbook, output_path: Path) -> None:
    # [ANCHOR:EXCEL_SAVE_WORKBOOK]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_sheet_preview(
    worksheet: Worksheet,
    *,
    header_row: int,
    preview_rows: int = 8,
) -> tuple[list[str], list[list[str]]]:
    # [ANCHOR:EXCEL_BUILD_SHEET_PREVIEW]
    headers: list[str] = []
    for column_index in range(1, worksheet.max_column + 1):
        raw_value = worksheet.cell(header_row, column_index).value
        normalized = normalize_text(str(raw_value or f"COL_{column_index}"))
        headers.append(build_column_label(column_index, normalized))

    rows: list[list[str]] = []
    data_end_row = min(worksheet.max_row, header_row + preview_rows)
    for row_index in range(header_row + 1, data_end_row + 1):
        row_values: list[str] = []
        for column_index in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row_index, column_index).value
            row_values.append("" if value is None else str(value))
        rows.append(row_values)
    return headers, rows
