from __future__ import annotations

from dataclasses import dataclass
import re
from openpyxl.worksheet.worksheet import Worksheet

from harmony_translate.preprocess import normalize_text


HEADER_NOTE_RE = re.compile(r"\bNOTE(S)?\b", re.IGNORECASE)
HEADER_EXCLUDE_RE = re.compile(
    r"SHOT\s*CODE|STARTING\s*FILE|TEAM|CLASSIFICATION|TIER|FRAME\s*COUNT|SECS|DURATION|ALL\s*TAGGED\s*ASSETS|^#$",
    re.IGNORECASE,
)
CODE_LIKE_RE = re.compile(
    r"^(HH\d{4}.*|[A-Z]{1,5}$|\d+(?:\.\d+)?|[A-Za-z0-9_.-]+(?:,\s*[A-Za-z0-9_.-]+)*)$"
)


@dataclass
class ColumnProfile:
    index: int
    header: str
    text_count: int
    numeric_count: int
    blank_count: int
    code_like_count: int
    unique_text_count: int
    average_length: float
    linebreak_ratio: float
    dedup_ratio: float
    character_count: int
    score: float


def profile_columns(
    worksheet: Worksheet,
    *,
    header_row: int,
    data_start_row: int,
    sample_size: int = 500,
) -> list[ColumnProfile]:
    # [ANCHOR:COLUMN_SELECTOR_PROFILE]
    profiles: list[ColumnProfile] = []
    data_end_row = min(worksheet.max_row, data_start_row + sample_size - 1)
    for column_index in range(1, worksheet.max_column + 1):
        header_value = worksheet.cell(header_row, column_index).value
        header = normalize_text(str(header_value or f"COL_{column_index}"))
        texts: list[str] = []
        numeric_count = 0
        blank_count = 0
        code_like_count = 0
        linebreak_count = 0
        for row_index in range(data_start_row, data_end_row + 1):
            value = worksheet.cell(row_index, column_index).value
            if value in (None, ""):
                blank_count += 1
                continue
            if isinstance(value, (int, float)):
                numeric_count += 1
                continue
            text = normalize_text(str(value))
            if not text:
                blank_count += 1
                continue
            texts.append(text)
            if CODE_LIKE_RE.fullmatch(text):
                code_like_count += 1
            if "\n" in text:
                linebreak_count += 1

        unique_count = len(set(texts))
        character_count = sum(len(text) for text in texts)
        average_length = (
            (sum(len(text) for text in texts) / len(texts)) if texts else 0.0
        )
        dedup_ratio = 1 - (unique_count / len(texts)) if texts else 0.0
        linebreak_ratio = linebreak_count / len(texts) if texts else 0.0
        score = score_column(
            header,
            texts,
            numeric_count,
            code_like_count,
            average_length,
            linebreak_ratio,
        )
        profiles.append(
            ColumnProfile(
                index=column_index,
                header=header,
                text_count=len(texts),
                numeric_count=numeric_count,
                blank_count=blank_count,
                code_like_count=code_like_count,
                unique_text_count=unique_count,
                average_length=average_length,
                linebreak_ratio=linebreak_ratio,
                dedup_ratio=dedup_ratio,
                character_count=character_count,
                score=score,
            )
        )
    return profiles


def score_column(
    header: str,
    texts: list[str],
    numeric_count: int,
    code_like_count: int,
    average_length: float,
    linebreak_ratio: float,
) -> float:
    # [ANCHOR:COLUMN_SELECTOR_SCORE]
    if HEADER_EXCLUDE_RE.search(header):
        return -100.0
    if not texts or len(texts) < 5:
        return -10.0
    code_ratio = code_like_count / max(len(texts), 1)
    if code_ratio >= 0.8:
        return -100.0
    score = 0.0
    if HEADER_NOTE_RE.search(header):
        score += 8.0
    if average_length >= 25:
        score += 4.0
    if average_length >= 60:
        score += 3.0
    if linebreak_ratio >= 0.1:
        score += 2.0
    if len(texts) >= max(20, numeric_count):
        score += 2.0
    if code_ratio >= 0.4:
        score -= 8.0
    return score


def select_translation_columns(profiles: list[ColumnProfile]) -> list[ColumnProfile]:
    # [ANCHOR:COLUMN_SELECTOR_SELECT]
    selected = [profile for profile in profiles if profile.score > 0]
    selected.sort(
        key=lambda profile: (profile.score, profile.text_count, profile.average_length),
        reverse=True,
    )
    return selected
